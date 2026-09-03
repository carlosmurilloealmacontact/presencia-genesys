"""
Extrae los turnos programados (Hora Inicio / Hora Fin) desde:
1. El archivo de Sistema de Punto compartido en red (Privado - ABS Sistema de Punto Ajustado.xlsx).
2. O un archivo semanal complementario (pasado por argumento o ubicado en data/turnos_semanales/).

Guarda/actualiza en SQLite (presencia_master.db y presencia.db) preservando el histórico acumulado.

Uso:
    python extract_turnos.py                           # Lee de red y luego de data/turnos_semanales/
    python extract_turnos.py "ruta/a/Turnos.xlsx"      # Lee un archivo especifico
"""

import sys
import os
import glob
from pathlib import Path
import openpyxl
import pandas as pd
import sqlite3

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

from db import get_connection, guardar_turnos
from jerarquia import load_cedula_a_bp
from config import CLOUD_EXPORT_PATH

RUTA_RED_DEFAULT = r"\\Co0000fs0001\01. Reporting\02. Sistema de Punto\Privado - ABS Sistema de Punto Ajustado.xlsx"
DIR_SEMANALES = Path(__file__).parent.parent / "data" / "turnos_semanales"


def procesar_archivo_red(ruta: str, cedula_a_bp: dict) -> list[dict]:
    """Procesa el formato tradicional de Sistema de Punto (hoja BD OPERACIÓN)."""
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    if "BD OPERACIÓN" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["BD OPERACIÓN"]

    COL_CODIGO = 1
    COL_FECHA = 7
    COL_HORA_INICIO = 13
    COL_HORA_FIN = 14

    rows = []
    vistos = set()
    for fila in ws.iter_rows(min_row=2, values_only=True):
        cedula = fila[COL_CODIGO]
        fecha = fila[COL_FECHA]
        hora_inicio = fila[COL_HORA_INICIO]
        hora_fin = fila[COL_HORA_FIN]
        if not cedula or not fecha or hora_inicio is None or hora_fin is None:
            continue

        cedula = str(cedula).strip()
        bp = cedula_a_bp.get(cedula)
        if not bp:
            continue

        fecha_str = fecha.strftime("%Y-%m-%d") if hasattr(fecha, "strftime") else str(fecha)[:10]
        clave = (bp, fecha_str)
        if clave in vistos:
            continue
        vistos.add(clave)

        h_ini = hora_inicio.strftime("%H:%M:%S") if hasattr(hora_inicio, "strftime") else str(hora_inicio)[:8]
        h_fin = hora_fin.strftime("%H:%M:%S") if hasattr(hora_fin, "strftime") else str(hora_fin)[:8]

        rows.append({
            "bp": bp,
            "fecha": fecha_str,
            "hora_inicio": h_ini,
            "hora_fin": h_fin,
        })

    wb.close()
    return rows


def procesar_archivo_semanal(ruta: str, cedula_a_bp: dict) -> list[dict]:
    """Procesa el formato exportado de turnos semanales (columnas Fecha, Documento, Turno_Ini, Turno_Fin, Novedad)."""
    df = pd.read_excel(ruta)
    # Normalizar nombres de columnas
    cols_map = {c.lower().strip(): c for c in df.columns}
    
    col_doc = cols_map.get("documento") or cols_map.get("cedula") or cols_map.get("codigocolaborador")
    col_fecha = cols_map.get("fecha")
    col_ini = cols_map.get("turno_ini") or cols_map.get("hora inicio") or cols_map.get("horainicio")
    col_fin = cols_map.get("turno_fin") or cols_map.get("hora fin") or cols_map.get("horafin")
    col_nov = cols_map.get("novedad")

    if not (col_doc and col_fecha and col_ini and col_fin):
        return []

    if col_nov:
        df = df[df[col_nov].astype(str).str.upper().str.strip().isin(["TUR", "TURNO", "TRB"])]

    df = df.dropna(subset=[col_doc, col_fecha, col_ini, col_fin])
    df["Documento_str"] = df[col_doc].astype(str).str.strip().str.split(".").str[0]
    df["bp"] = df["Documento_str"].map(cedula_a_bp)
    df = df[df["bp"].notna()].copy()

    df["fecha_str"] = pd.to_datetime(df[col_fecha]).dt.strftime("%Y-%m-%d")
    df["h_ini"] = pd.to_datetime(df[col_ini].astype(str), format="%H:%M:%S", errors="coerce").dt.strftime("%H:%M:%S")
    df["h_fin"] = pd.to_datetime(df[col_fin].astype(str), format="%H:%M:%S", errors="coerce").dt.strftime("%H:%M:%S")

    df = df.dropna(subset=["h_ini", "h_fin"])
    df = df.drop_duplicates(subset=["bp", "fecha_str"], keep="last")

    rows = [
        {
            "bp": str(r["bp"]).strip(),
            "fecha": str(r["fecha_str"]).strip(),
            "hora_inicio": str(r["h_ini"]).strip(),
            "hora_fin": str(r["h_fin"]).strip(),
        }
        for _, r in df.iterrows()
    ]
    return rows


def procesar_archivo(ruta: str, cedula_a_bp: dict) -> list[dict]:
    """Detecta el formato del archivo y extrae los turnos."""
    print(f"Leyendo {ruta} ...")
    try:
        # Intentar primero como archivo de red BD OPERACIÓN
        rows = procesar_archivo_red(ruta, cedula_a_bp)
        if rows:
            print(f"  Formato BD OPERACIÓN detectado: {len(rows)} turnos leídos.")
            return rows
    except Exception:
        pass

    try:
        # Intentar como formato semanal exportado
        rows = procesar_archivo_semanal(ruta, cedula_a_bp)
        if rows:
            print(f"  Formato Semanal detectado: {len(rows)} turnos leídos.")
            return rows
    except Exception as e:
        print(f"  Error procesando {ruta}: {e}")

    return []


def run(archivo_especifico: str = None):
    print("Cargando puente cedula -> BP desde el sheet Base...")
    cedula_a_bp = load_cedula_a_bp()
    print(f"  {len(cedula_a_bp)} cedulas mapeadas.")

    todos_los_turnos = []

    if archivo_especifico:
        if os.path.exists(archivo_especifico):
            todos_los_turnos.extend(procesar_archivo(archivo_especifico, cedula_a_bp))
        else:
            print(f"ERROR: No se encontró el archivo especificado: {archivo_especifico}")
            return
    else:
        # 1. Intentar archivo de red si está disponible
        if os.path.exists(RUTA_RED_DEFAULT):
            todos_los_turnos.extend(procesar_archivo(RUTA_RED_DEFAULT, cedula_a_bp))
        else:
            print(f"Aviso: Ruta de red no accesible ({RUTA_RED_DEFAULT}).")

        # 2. Buscar archivos complementarios en data/turnos_semanales/
        if DIR_SEMANALES.exists():
            for f in sorted(DIR_SEMANALES.glob("*.xlsx")):
                print(f"Procesando archivo complementario: {f.name}")
                todos_los_turnos.extend(procesar_archivo(str(f), cedula_a_bp))

    if not todos_los_turnos:
        print("No se encontraron turnos para guardar.")
        return

    # Guardar en presencia_master.db
    conn = get_connection()
    guardar_turnos(conn, todos_los_turnos)
    total_master = conn.execute("SELECT COUNT(*), MIN(fecha), MAX(fecha) FROM turnos").fetchone()
    conn.close()
    print(f"Guardado en presencia_master.db: {total_master[0]} registros acumulados ({total_master[1]} -> {total_master[2]}).")

    # Sincronizar con presencia.db (usado por el visor local y Streamlit Cloud)
    cloud_db = Path(__file__).parent / CLOUD_EXPORT_PATH
    if cloud_db.exists():
        conn_cloud = sqlite3.connect(cloud_db)
        guardar_turnos(conn_cloud, todos_los_turnos)
        total_cloud = conn_cloud.execute("SELECT COUNT(*), MIN(fecha), MAX(fecha) FROM turnos").fetchone()
        conn_cloud.close()
        print(f"Sincronizado en presencia.db: {total_cloud[0]} registros acumulados ({total_cloud[1]} -> {total_cloud[2]}).")


if __name__ == "__main__":
    arg = sys.argv[1] if len(sys.argv) > 1 else None
    run(arg)
