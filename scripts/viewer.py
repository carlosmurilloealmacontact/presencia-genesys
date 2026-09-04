# Radar Genesys — Pausas y Adherencia de Turno.
# Uso: streamlit run viewer.py
import numpy as np

import sqlite3
from io import BytesIO
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.utils import get_column_letter

from config import DB_PATH
import os
from live_engine import render_tab_en_vivo

st.set_page_config(page_title="Radar Genesys", layout="wide")

PALETA_ESTADOS = px.colors.qualitative.Alphabet + px.colors.qualitative.Dark24

# Tarjetas de cumplimiento.
# tipo "graduada": 0 usado = 0%; 0 < usado <= meta = 100%; usado > meta baja
#   suave (piso(100*meta/usado)) - no es binario, entre mas se pase mas cae.
# tipo "conteo": sin meta, solo se cuenta cuantas veces se uso (Lunch).
# "presence" es una lista: Dialogo une PCA-Dialogo + Dialogo Diario/4DX en una
#   sola meta de 15 min/dia (un agente usa una u otra, no las dos).
# unidad: "dia" o "semana" (CDR es semanal).
CARDS = [
    {"key": "descanso",    "label": "Descanso",               "presence": ["Break"],                                    "meta": 30, "unidad": "dia",    "tipo": "graduada"},
    {"key": "pre_pausa",   "label": "Ocupado: Pre Pausa",      "presence": ["Pre Pausa"],                                "meta": 60, "unidad": "dia",    "tipo": "graduada"},
    {"key": "bano",        "label": "Ausente: Baño",          "presence": ["Baño"],                                     "meta": 5,  "unidad": "dia",    "tipo": "graduada"},
    {"key": "dialogo",     "label": "Reunión: Diálogo",       "presence": ["PCA- Diálogo", "Diálogo Diario / 4DX"],     "meta": 15, "unidad": "dia",    "tipo": "graduada"},
    {"key": "cdr",         "label": "Reunión: CDR",            "presence": ["CDR"],                                      "meta": 30, "unidad": "semana", "tipo": "graduada"},
    {"key": "refuerzo",    "label": "Capacitación: Refuerzo", "presence": ["Refuerzo Semanal"],                         "meta": 60, "unidad": "semana", "tipo": "graduada"},
    {"key": "autogestion", "label": "Autogestión",            "presence": ["Autogestión"],                              "meta": 30, "unidad": "semana", "tipo": "graduada"},
    {"key": "feedback",    "label": "Feedback",               "presence": ["Feedback", "PCA - Feedback"],               "meta": 30, "unidad": "semana", "tipo": "graduada"},
    {"key": "cursos",      "label": "Cursos Adicionales",     "presence": ["Cursos Adicionales"],                       "meta": 60, "unidad": "semana", "tipo": "graduada"},
    {"key": "lunch",       "label": "Comida: Lunch",           "presence": ["Lunch"],                                    "meta": None, "unidad": "dia",  "tipo": "conteo"},
]

ESTADOS_SISTEMA = {"Offline", "Available", "Conectado", "On Queue"}

PAUSAS_NO_AUTORIZADAS = {
    "Lunch",
    "Refeição (sólo BR)",
    "Ampliado",
    "Reunión Equipo",
    "LATAM News",
    "Casos borde (sólo LAE)",
    "Boletín RT (sólo LAE)",
    "Mediación entidades (sólo LAE)",
    "Corresponsables (sólo LAE)",
    "Gestión sin Contacto",
    "Problemas de Acceso",
}

def servicio_autorizado_casos_bo(servicio: str) -> bool:
    """Casos Backoffice está autorizado para servicios BO y células de Chat/Redes autorizadas."""
    s = (servicio or "").upper()
    if s.startswith("BO ") or s.startswith("BO_") or "BACKOFFICE" in s:
        return True
    celulas = (
        "RRSS AMC",
        "CHAT AGENCIAS ESP",
        "AG CORPORATE CHAT",
        "SPEECH",
        "AG CHECK IN",
        "AG CELULA REMISION",
        "CARGO BOOKING",
        "RRSS AMC ING",
        "RRSS PORT AMC",
        "ANTIFRAUDE",
    )
    return any(c in s for c in celulas)

# Asesores de Backoffice / Células especializadas con alto uso legítimo de Available
AGENTES_AUTORIZADOS_AVAILABLE_DEFAULT = [
    "4452121 - Galeano Arboleda Cristian Alfredo",
    "4509968 - Macias Cardenas Mateo",
    "4810816 - Ospina Porras Juan Felipe",
    "4856439 - Morales Albarracin Juan Esteban",
    "4856439 - Juan Esteban Morales Albarracin",
    "4770490 - Martinez Benitez Carol Dariena",
    "4690319 - Chaverra Aguirre Mariana",
    "4450680 - Higuita Guerra Yadi Lorena",
    "4758833 - David Felipe Pardo Anzola",
    "4818871 - Toro Munera Juan Manuel",
    "4718011 - Perez Hincapie Yolanda",
    "4614395 - Restrepo Uribe Emanuel",
    "4729574 - Darly Dayana Guzman Jimenez",
    "4819011 - Rojas Rodriguez Mairene Beatriz",
    "4790270 - Tobon Cardenas Maria Isabel",
    "4853852 - Carrascal Gonzalez Luciana",
    "4636625 - Mosquera Vergara Jhoseline Dayana",
    "4822621 - Cano Gaviria Ana Maria",
    "4196581 - QUINTERO BETANCUR KELLY GEOVANNA",
    "4496451 - Rutberling Delcarmen Rojas Anderson",
]


def meta_texto(card: dict) -> str:
    if card["tipo"] == "conteo":
        return "Sin meta — solo conteo"
    unidad_txt = "día" if card["unidad"] == "dia" else "semana"
    return f"Meta: {card['meta']} min/{unidad_txt}"


def color_pct(pct: float) -> str:
    if pct < 70:
        return "#e24b4a"
    if pct < 90:
        return "#eda100"
    return "#1baf7a"


def color_uso_turno(pct: float) -> str:
    """Semáforo estricto para aprovechamiento de turno: <85% rojo, 85-95% amarillo, >=95% verde."""
    if pct < 85:
        return "#e24b4a"
    if pct < 95:
        return "#eda100"
    return "#1baf7a"


def color_fuga(pct: float) -> str:
    """Semáforo para fuga en conexión (excesos e improductivas): <=3% verde, <=6% amarillo, >6% rojo."""
    if pct <= 3.0:
        return "#1baf7a"
    if pct <= 6.0:
        return "#eda100"
    return "#e24b4a"


def estilo_prod(val):
    if pd.isna(val):
        return ""
    color = "#1baf7a" if val >= 85 else ("#eda100" if val >= 75 else "#e24b4a")
    return f"background-color: {color}22; color: {color}; font-weight: 600;"


def estilo_aut(val):
    if pd.isna(val):
        return ""
    return "background-color: #378ADD22; color: #378ADD; font-weight: 600;"


def estilo_micro(val):
    """Estilo para micro-estados (<= 15 seg): alerta si es recurrente."""
    if pd.isna(val) or val == 0:
        return ""
    if val >= 20:
        color = "#e24b4a"
    elif val >= 5:
        color = "#eda100"
    else:
        color = "#378ADD"
    return f"background-color: {color}22; color: {color}; font-weight: 600;"


def servicio_tiene_prepausa(servicio: str) -> bool:
    """Pre Pausa aplica a canales WPP/Chat/RRSS, excepto CHAT AGENCIAS ESP."""
    s = (servicio or "").upper()
    if "CHAT AGENCIAS ESP" in s:
        return False
    return any(k in s for k in ("WPP", "CHAT", "RRSS"))


def calcular_componentes_100(
    df_sub: pd.DataFrame,
    servicio: str = "",
    penalizar_available: bool = True,
    excluidos_available: set = None,
):
    """Calcula minutos productivos, pausas autorizadas y excesos sobre el tiempo conectado (100%)."""
    if excluidos_available is None:
        excluidos_available = set()

    conectados_df = df_sub[df_sub["presence_label"] != "Offline"].copy()
    t_con = float(conectados_df["duracion_min"].sum()) if not conectados_df.empty else 0.0
    if t_con <= 0:
        return 0.0, 0.0, 0.0, 0.0, {}, {}, {}

    dt = pd.to_datetime(conectados_df["fecha"])
    iso = dt.dt.isocalendar()
    conectados_df["semana"] = iso.year.astype(str) + "-W" + iso.week.astype(str).str.zfill(2)

    # 1. Diario (Break, Baño, Diálogo, Pre Pausa, no autorizadas)
    diario = (
        conectados_df.groupby(["agente", "fecha", "semana", "servicio", "presence_label"])["duracion_min"]
        .sum()
        .unstack(fill_value=0.0)
        .reset_index()
    )

    tot_aut = 0.0
    tot_exc = 0.0
    detalle_aut = {}
    detalle_exc = {}
    detalle_prod = {}

    for _, row in diario.iterrows():
        agente = row["agente"]
        serv = row["servicio"] if "servicio" in row and pd.notna(row["servicio"]) else servicio
        tiene_pre = servicio_tiene_prepausa(serv)
        meta_pre = 60.0 if tiene_pre else 0.0

        bano = row.get("Baño", 0.0)
        b_aut = min(bano, 5.0)
        b_exc = max(0.0, bano - 5.0)

        brk = row.get("Break", 0.0)
        brk_aut = min(brk, 30.0)
        brk_exc = max(0.0, brk - 30.0)

        dia = row.get("Diálogo Diario / 4DX", 0.0) + row.get("PCA- Diálogo", 0.0)
        dia_aut = min(dia, 15.0)
        dia_exc = max(0.0, dia - 15.0)

        pre = row.get("Pre Pausa", 0.0)
        pre_aut = min(pre, meta_pre)
        pre_exc = max(0.0, pre - meta_pre)

        # Determinar pausas no autorizadas
        pausas_no_aut = set(PAUSAS_NO_AUTORIZADAS)
        if penalizar_available and (agente not in excluidos_available):
            pausas_no_aut.update({"Available", "Conectado"})
        if not servicio_autorizado_casos_bo(serv):
            pausas_no_aut.add("Casos Backoffice")

        tot_no_aut = 0.0
        for p_na in pausas_no_aut:
            v_na = row.get(p_na, 0.0)
            if v_na > 0:
                tot_no_aut += v_na
                detalle_exc[f"{p_na} (No aut.)"] = detalle_exc.get(f"{p_na} (No aut.)", 0.0) + v_na

        tot_aut += (b_aut + brk_aut + dia_aut + pre_aut)
        tot_exc += (b_exc + brk_exc + dia_exc + pre_exc + tot_no_aut)

        detalle_aut["Descanso (Break)"] = detalle_aut.get("Descanso (Break)", 0.0) + brk_aut
        detalle_aut["Baño"] = detalle_aut.get("Baño", 0.0) + b_aut
        detalle_aut["Diálogo / 4DX"] = detalle_aut.get("Diálogo / 4DX", 0.0) + dia_aut
        if tiene_pre and pre_aut > 0:
            detalle_aut["Pre Pausa"] = detalle_aut.get("Pre Pausa", 0.0) + pre_aut

        if brk_exc > 0:
            detalle_exc["Exceso Descanso"] = detalle_exc.get("Exceso Descanso", 0.0) + brk_exc
        if b_exc > 0:
            detalle_exc["Exceso Baño"] = detalle_exc.get("Exceso Baño", 0.0) + b_exc
        if dia_exc > 0:
            detalle_exc["Exceso Diálogo"] = detalle_exc.get("Exceso Diálogo", 0.0) + dia_exc
        if pre_exc > 0:
            detalle_exc["Exceso Pre Pausa"] = detalle_exc.get("Exceso Pre Pausa", 0.0) + pre_exc

        # Available si está en lista blanca
        if not (penalizar_available and (agente not in excluidos_available)):
            for p_av in ("Available", "Conectado"):
                v_av = row.get(p_av, 0.0)
                if v_av > 0:
                    detalle_prod[p_av] = detalle_prod.get(p_av, 0.0) + v_av

        # Casos Backoffice si está en servicio autorizado
        if servicio_autorizado_casos_bo(serv):
            v_bo = row.get("Casos Backoffice", 0.0)
            if v_bo > 0:
                detalle_prod["Casos Backoffice"] = detalle_prod.get("Casos Backoffice", 0.0) + v_bo

    # 2. Semanal (CDR 30m, Refuerzo 60m, Autogestión 30m, Feedback 30m, Cursos 60m)
    semanal = (
        conectados_df.groupby(["agente", "semana", "presence_label"])["duracion_min"]
        .sum()
        .unstack(fill_value=0.0)
        .reset_index()
    )

    for _, row in semanal.iterrows():
        cdr = row.get("CDR", 0.0)
        cdr_aut = min(cdr, 30.0)
        cdr_exc = max(0.0, cdr - 30.0)

        ref = row.get("Refuerzo Semanal", 0.0)
        ref_aut = min(ref, 60.0)
        ref_exc = max(0.0, ref - 60.0)

        auto = row.get("Autogestión", 0.0)
        auto_aut = min(auto, 30.0)
        auto_exc = max(0.0, auto - 30.0)

        feed = row.get("Feedback", 0.0) + row.get("PCA - Feedback", 0.0)
        feed_exc = max(0.0, feed - 30.0)
        feed_prod = min(feed, 30.0)

        cur = row.get("Cursos Adicionales", 0.0)
        cur_exc = max(0.0, cur - 60.0)
        cur_prod = min(cur, 60.0)

        tot_aut += (cdr_aut + ref_aut + auto_aut)
        tot_exc += (cdr_exc + ref_exc + auto_exc + feed_exc + cur_exc)

        if cdr_aut > 0:
            detalle_aut["CDR"] = detalle_aut.get("CDR", 0.0) + cdr_aut
        if ref_aut > 0:
            detalle_aut["Refuerzo Semanal"] = detalle_aut.get("Refuerzo Semanal", 0.0) + ref_aut
        if auto_aut > 0:
            detalle_aut["Autogestión"] = detalle_aut.get("Autogestión", 0.0) + auto_aut

        if cdr_exc > 0:
            detalle_exc["Exceso CDR"] = detalle_exc.get("Exceso CDR", 0.0) + cdr_exc
        if ref_exc > 0:
            detalle_exc["Exceso Refuerzo"] = detalle_exc.get("Exceso Refuerzo", 0.0) + ref_exc
        if auto_exc > 0:
            detalle_exc["Exceso Autogestión"] = detalle_exc.get("Exceso Autogestión", 0.0) + auto_exc
        if feed_exc > 0:
            detalle_exc["Exceso Feedback (>30m/sem)"] = detalle_exc.get("Exceso Feedback (>30m/sem)", 0.0) + feed_exc
        if cur_exc > 0:
            detalle_exc["Exceso Cursos (>60m/sem)"] = detalle_exc.get("Exceso Cursos (>60m/sem)", 0.0) + cur_exc

        if feed_prod > 0:
            detalle_prod["Feedback"] = detalle_prod.get("Feedback", 0.0) + feed_prod
        if cur_prod > 0:
            detalle_prod["Cursos Adicionales"] = detalle_prod.get("Cursos Adicionales", 0.0) + cur_prod

    # Registrar el resto de presencias productivas directas (On Queue, etc.)
    presencias_controladas = {
        "Break", "Baño", "Diálogo Diario / 4DX", "PCA- Diálogo", "Lunch", "Pre Pausa",
        "CDR", "Refuerzo Semanal", "Autogestión", "Feedback", "PCA - Feedback", "Cursos Adicionales",
        "Casos Backoffice", "Available", "Conectado", "Offline"
    }.union(PAUSAS_NO_AUTORIZADAS)

    for col in conectados_df["presence_label"].unique():
        if col not in presencias_controladas:
            val = float(conectados_df[conectados_df["presence_label"] == col]["duracion_min"].sum())
            if val > 0:
                detalle_prod[col] = val

    prod_min = max(0.0, t_con - tot_aut - tot_exc)
    return prod_min, tot_aut, tot_exc, t_con, detalle_prod, detalle_aut, detalle_exc

def render_distribucion_100(prod_min: float, aut_min: float, exc_min: float, t_con_min: float, titulo: str = "Distribución del 100% del Tiempo Conectado"):
    """Renderiza una barra horizontal apilada que suma exactamente el 100% del tiempo conectado."""
    if t_con_min <= 0:
        return
    pct_prod = round(prod_min / t_con_min * 100.0, 1)
    pct_aut = round(aut_min / t_con_min * 100.0, 1)
    pct_exc = round(exc_min / t_con_min * 100.0, 1)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=["Conexión"], x=[pct_prod], name="Productivo", orientation="h",
        marker=dict(color="#1baf7a"),
        text=[f"Productivo<br>{pct_prod}% ({prod_min:.0f}m)"], textposition="inside", insidetextanchor="middle",
        hovertemplate="<b>Productivo</b>: %{x:.1f}% (" + f"{prod_min:.1f} min)<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        y=["Conexión"], x=[pct_aut], name="Pausas Autorizadas", orientation="h",
        marker=dict(color="#378ADD"),
        text=[f"Pausas Aut.<br>{pct_aut}% ({aut_min:.0f}m)"] if pct_aut >= 6 else [""], textposition="inside", insidetextanchor="middle",
        hovertemplate="<b>Pausas Autorizadas</b>: %{x:.1f}% (" + f"{aut_min:.1f} min)<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        y=["Conexión"], x=[pct_exc], name="Excesos / Fuga", orientation="h",
        marker=dict(color="#e24b4a"),
        text=[f"Excesos/Fuga<br>{pct_exc}% ({exc_min:.0f}m)"] if pct_exc >= 6 else [""], textposition="inside", insidetextanchor="middle",
        hovertemplate="<b>Excesos / Fuga</b>: %{x:.1f}% (" + f"{exc_min:.1f} min)<extra></extra>",
    ))
    fig.update_layout(
        barmode="stack",
        height=130,
        margin=dict(l=10, r=10, t=32, b=10),
        xaxis=dict(showgrid=False, showticklabels=True, range=[0, 100], ticksuffix="%"),
        yaxis=dict(showticklabels=False),
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        title=dict(text=f"<b>{titulo}</b> — Total: {t_con_min/60.0:.1f} hrs ({t_con_min:.0f} min = 100%)", font=dict(size=13)),
    )
    st.plotly_chart(fig, use_container_width=True)


def sumar_presencias(pivot: pd.DataFrame, presence_list: list[str]) -> pd.Series:
    cols = [p for p in presence_list if p in pivot.columns]
    if not cols:
        return pd.Series(0.0, index=pivot.index)
    return pivot[cols].sum(axis=1)


def score_graduado(usado: pd.Series, meta: float) -> pd.Series:
    """0 usado -> 0%; 0<usado<=meta -> 100%; usado>meta -> piso(100*meta/usado)."""
    score = pd.Series(0.0, index=usado.index)
    cumple_pleno = (usado > 0) & (usado <= meta)
    score[cumple_pleno] = 100.0
    excedido = usado > meta
    score[excedido] = (100 * meta / usado[excedido]).apply(lambda x: float(int(x)))
    return score


# ── Datos ────────────────────────────────────────────────────────────────

@st.cache_data(ttl=60)
def cargar_rango_fechas() -> tuple[str, str] | None:
    db_path = Path(__file__).parent / DB_PATH
    if not db_path.exists():
        return None
    with sqlite3.connect(db_path) as conn:
        row = conn.execute("SELECT MIN(fecha), MAX(fecha) FROM segments").fetchone()
    return row if row and row[0] else None


@st.cache_data(ttl=60)
def cargar_rango(fecha_min: str, fecha_max: str) -> pd.DataFrame:
    db_path = Path(__file__).parent / DB_PATH
    with sqlite3.connect(db_path) as conn:
        df = pd.read_sql_query(
            "SELECT * FROM segments WHERE fecha BETWEEN ? AND ?", conn, params=(fecha_min, fecha_max)
        )
    if "cargo" in df.columns:
        df = df[df["cargo"].str.upper().str.contains("ASESOR", na=False)]
    return df


@st.cache_data(ttl=60)
def cargar_turnos(fecha_min: str, fecha_max: str) -> pd.DataFrame:
    db_path = Path(__file__).parent / DB_PATH
    with sqlite3.connect(db_path) as conn:
        try:
            df = pd.read_sql_query(
                "SELECT * FROM turnos WHERE fecha BETWEEN ? AND ?", conn, params=(fecha_min, fecha_max)
            )
        except pd.errors.DatabaseError:
            df = pd.DataFrame(columns=["bp", "fecha", "hora_inicio", "hora_fin"])
    return df


@st.cache_data(ttl=300)
def cargar_color_map() -> dict[str, str]:
    """Color fijo por estado sobre el universo completo ya visto en la base."""
    db_path = Path(__file__).parent / DB_PATH
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute("SELECT DISTINCT presence_label FROM segments ORDER BY presence_label").fetchall()
    universo = [r[0] for r in rows]
    return {estado: PALETA_ESTADOS[i % len(PALETA_ESTADOS)] for i, estado in enumerate(universo)}


def opciones_validas(serie: pd.Series) -> list[str]:
    return sorted(v for v in serie.unique() if v)


# ── Calculo de cumplimiento ─────────────────────────────────────────────

def calcular_cumplimiento(df: pd.DataFrame) -> pd.DataFrame:
    """
    Una fila por agente. Para tarjetas "graduada": promedio del score diario
    (o semanal) de cada dia/semana trabajado — no es un % de dias que pasaron
    un corte binario, cada dia tiene su propio score 0-100 (ver score_graduado).
    Para "conteo" (Lunch): total de veces usado en el rango, sin promediar.
    Denominador de dias/semanas = cuando el agente tuvo CUALQUIER tramo de
    presencia, no dias calendario, para no penalizar dias libres que no vemos.
    """
    info_agente = df.drop_duplicates("agente_id").set_index("agente_id")[
        ["agente", "servicio", "jefe_inmediato", "coordinador"]
    ]

    df = df.copy()
    dt = pd.to_datetime(df["fecha"])
    iso = dt.dt.isocalendar()
    df["semana"] = iso.year.astype(str) + "-W" + iso.week.astype(str).str.zfill(2)

    resultado = pd.DataFrame(index=info_agente.index)

    diario = df.groupby(["agente_id", "fecha", "presence_label"])["duracion_min"].sum().reset_index()
    pivot_diario = diario.pivot_table(index=["agente_id", "fecha"], columns="presence_label", values="duracion_min", fill_value=0)
    dias_trabajados = pivot_diario.groupby("agente_id").size()
    resultado["dias_trabajados"] = dias_trabajados

    semanal = df.groupby(["agente_id", "semana", "presence_label"])["duracion_min"].sum().reset_index()
    pivot_semanal = semanal.pivot_table(index=["agente_id", "semana"], columns="presence_label", values="duracion_min", fill_value=0)
    semanas_trabajadas = pivot_semanal.groupby("agente_id").size()
    resultado["semanas_trabajadas"] = semanas_trabajadas

    for c in CARDS:
        if c["tipo"] == "conteo":
            conteo = df[df["presence_label"].isin(c["presence"])].groupby("agente_id").size()
            resultado[c["key"]] = conteo.reindex(resultado.index, fill_value=0).astype(int)
            continue
        pivot = pivot_diario if c["unidad"] == "dia" else pivot_semanal
        usado = sumar_presencias(pivot, c["presence"])
        score = score_graduado(usado, c["meta"])
        resultado[c["key"]] = score.groupby("agente_id").mean().round(1)

    return resultado.join(info_agente)


def cumplimiento_global(df: pd.DataFrame) -> dict[str, float]:
    """Promedio pooled (no promedio de promedios) por tarjeta, para las tarjetas KPI de arriba."""
    df = df.copy()
    dt = pd.to_datetime(df["fecha"])
    iso = dt.dt.isocalendar()
    df["semana"] = iso.year.astype(str) + "-W" + iso.week.astype(str).str.zfill(2)

    diario = df.groupby(["agente_id", "fecha", "presence_label"])["duracion_min"].sum().reset_index()
    pivot_diario = diario.pivot_table(index=["agente_id", "fecha"], columns="presence_label", values="duracion_min", fill_value=0)

    semanal = df.groupby(["agente_id", "semana", "presence_label"])["duracion_min"].sum().reset_index()
    pivot_semanal = semanal.pivot_table(index=["agente_id", "semana"], columns="presence_label", values="duracion_min", fill_value=0)

    globales = {}
    for c in CARDS:
        if c["tipo"] == "conteo":
            globales[c["key"]] = int(df[df["presence_label"].isin(c["presence"])].shape[0])
            continue
        pivot = pivot_diario if c["unidad"] == "dia" else pivot_semanal
        if pivot.empty:
            globales[c["key"]] = 0.0
            continue
        usado = sumar_presencias(pivot, c["presence"])
        score = score_graduado(usado, c["meta"])
        globales[c["key"]] = round(score.mean(), 1)
    return globales


def calcular_adherencia_horario(vista: pd.DataFrame, turnos: pd.DataFrame) -> pd.DataFrame:
    """
    % del turno programado (Hora Inicio/Fin, archivo de Sistema de Punto) en
    que el agente estuvo en un estado distinto de Offline en Genesys. Cruza
    por BP -> agente_id (el archivo de turnos identifica por BP).
    """
    columnas_vacias = ["horario", "dias_con_turno"]
    if turnos.empty or vista.empty:
        return pd.DataFrame(columns=["agente_id"] + columnas_vacias).set_index("agente_id")

    vista = vista.copy()
    vista["bp"] = vista["agente"].str.split(" - ").str[0].str.strip()
    bp_a_agente_id = vista.groupby("bp")["agente_id"].agg(lambda s: s.mode().iat[0])

    productivos = vista[vista["presence_label"] != "Offline"].copy()
    productivos["inicio"] = pd.to_datetime(productivos["inicio"])
    productivos["fin"] = pd.to_datetime(productivos["fin"]).fillna(productivos["inicio"] + pd.Timedelta(minutes=1))
    segmentos_por_dia = {
        clave: list(zip(grupo["inicio"], grupo["fin"]))
        for clave, grupo in productivos.groupby(["agente_id", "fecha"])
    }

    acumulado = {}
    for row in turnos.itertuples(index=False):
        agente_id = bp_a_agente_id.get(row.bp)
        if agente_id is None:
            continue

        fecha_dt = pd.Timestamp(row.fecha)
        inicio_turno = fecha_dt + pd.to_timedelta(row.hora_inicio)
        fin_turno = fecha_dt + pd.to_timedelta(row.hora_fin)
        if fin_turno <= inicio_turno:
            fin_turno += pd.Timedelta(days=1)  # turno nocturno que cruza medianoche
        duracion_turno = (fin_turno - inicio_turno).total_seconds() / 60
        if duracion_turno <= 0:
            continue

        overlap_min = 0.0
        fecha_siguiente = (fecha_dt + pd.Timedelta(days=1)).strftime("%Y-%m-%d")
        for clave in ((agente_id, row.fecha), (agente_id, fecha_siguiente)):
            for seg_inicio, seg_fin in segmentos_por_dia.get(clave, []):
                ini = max(seg_inicio, inicio_turno)
                fin = min(seg_fin, fin_turno)
                if fin > ini:
                    overlap_min += (fin - ini).total_seconds() / 60

        acc = acumulado.setdefault(agente_id, {"overlap": 0.0, "programado": 0.0, "dias": 0})
        acc["overlap"] += min(overlap_min, duracion_turno)
        acc["programado"] += duracion_turno
        acc["dias"] += 1

    filas = [
        {
            "agente_id": agente_id,
            "horario": round(100 * acc["overlap"] / acc["programado"], 1) if acc["programado"] > 0 else float("nan"),
            "dias_con_turno": acc["dias"],
        }
        for agente_id, acc in acumulado.items()
    ]
    if not filas:
        return pd.DataFrame(columns=["agente_id"] + columnas_vacias).set_index("agente_id")
    return pd.DataFrame(filas).set_index("agente_id")


def calcular_uso_turno(
    vista: pd.DataFrame,
    turnos: pd.DataFrame,
    penalizar_available: bool = True,
    excluidos_available: set = None,
):
    """Calcula el aprovechamiento descontando excesos de pausas y pausas no autorizadas."""
    if excluidos_available is None:
        excluidos_available = set()

    columnas_vacias = ["pct_fuga_con", "pct_productivo", "pct_pausa_aut", "exceso_prom_min", "t_conectado_hrs"]
    if vista.empty:
        return (
            pd.DataFrame(columns=["agente_id"] + columnas_vacias).set_index("agente_id"),
            {"pct_fuga_con": 0.0, "pct_productivo": 0.0, "pct_pausa_aut": 0.0, "exceso_prom_min": 0.0, "t_conectado_hrs": 0.0},
            {},
        )

    turnos_validos = turnos.dropna(subset=["hora_inicio", "hora_fin"]).copy()
    if not turnos_validos.empty:
        turnos_validos["inicio_dt"] = pd.to_datetime(
            turnos_validos["fecha"] + " " + turnos_validos["hora_inicio"]
        )
        turnos_validos["fin_dt"] = pd.to_datetime(
            turnos_validos["fecha"] + " " + turnos_validos["hora_fin"]
        )
        turnos_validos.loc[
            turnos_validos["fin_dt"] < turnos_validos["inicio_dt"], "fin_dt"
        ] += pd.Timedelta(days=1)
        turnos_validos["duracion_min"] = (
            turnos_validos["fin_dt"] - turnos_validos["inicio_dt"]
        ).dt.total_seconds() / 60.0
        mapa_turnos = turnos_validos.set_index(["bp", "fecha"])["duracion_min"].to_dict()
    else:
        mapa_turnos = {}

    vista_temp = vista.copy()
    vista_temp["bp"] = vista_temp["agente"].str.split(" - ").str[0].str.strip()
    dt = pd.to_datetime(vista_temp["fecha"])
    iso = dt.dt.isocalendar()
    vista_temp["semana"] = iso.year.astype(str) + "-W" + iso.week.astype(str).str.zfill(2)

    df_con = vista_temp[vista_temp["presence_label"] != "Offline"].copy()
    conectados = df_con.groupby(["agente_id", "fecha"])["duracion_min"].sum().to_dict()

    # 1. Diario (Break, Baño, Diálogo, Pre Pausa, No autorizadas)
    diario = (
        df_con.groupby(["agente_id", "agente", "bp", "servicio", "fecha", "semana", "presence_label"])["duracion_min"]
        .sum()
        .unstack(fill_value=0.0)
        .reset_index()
    )

    filas_dias = []
    for _, row in diario.iterrows():
        agente_id = row["agente_id"]
        agente = row["agente"]
        bp = row["bp"]
        servicio = row["servicio"]
        fecha = row["fecha"]
        semana = row["semana"]

        duracion_turno = mapa_turnos.get((bp, fecha))
        t_conectado = conectados.get((agente_id, fecha), 0.0)
        if not duracion_turno or duracion_turno <= 0:
            duracion_turno = t_conectado if t_conectado > 0 else 480.0

        bano = row.get("Baño", 0.0)
        b_aut = min(bano, 5.0)
        b_exc = max(0.0, bano - 5.0)

        descanso = row.get("Break", 0.0)
        brk_aut = min(descanso, 30.0)
        brk_exc = max(0.0, descanso - 30.0)

        dialogo = row.get("Diálogo Diario / 4DX", 0.0) + row.get("PCA- Diálogo", 0.0)
        dia_aut = min(dialogo, 15.0)
        dia_exc = max(0.0, dialogo - 15.0)

        pre_pausa = row.get("Pre Pausa", 0.0)
        meta_pre = 60.0 if servicio_tiene_prepausa(servicio) else 0.0
        pre_aut = min(pre_pausa, meta_pre)
        pre_exc = max(0.0, pre_pausa - meta_pre)

        # Pausas no autorizadas
        penalizar_este = penalizar_available and (agente not in excluidos_available)
        no_aut_lista = set(PAUSAS_NO_AUTORIZADAS)
        if penalizar_este:
            no_aut_lista.update({"Available", "Conectado"})
        if not servicio_autorizado_casos_bo(servicio):
            no_aut_lista.add("Casos Backoffice")

        no_aut_min = sum(row.get(col, 0.0) for col in no_aut_lista)
        exc_dia = b_exc + brk_exc + dia_exc + pre_exc + no_aut_min
        aut_dia = b_aut + brk_aut + dia_aut + pre_aut

        filas_dias.append({
            "agente_id": agente_id,
            "agente": agente,
            "fecha": fecha,
            "semana": semana,
            "t_conectado_min": t_conectado,
            "exc_dia": exc_dia,
            "aut_dia": aut_dia,
        })

    df_dias = pd.DataFrame(filas_dias)

    # 2. Semanal (CDR 30m, Refuerzo 60m, Autogestión 30m, Feedback 30m, Cursos 60m)
    semanal = (
        df_con.groupby(["agente_id", "semana", "presence_label"])["duracion_min"]
        .sum()
        .unstack(fill_value=0.0)
        .reset_index()
    )

    filas_sem = []
    for _, row in semanal.iterrows():
        agente_id = row["agente_id"]
        semana = row["semana"]

        cdr = row.get("CDR", 0.0)
        cdr_aut = min(cdr, 30.0)
        cdr_exc = max(0.0, cdr - 30.0)

        ref = row.get("Refuerzo Semanal", 0.0)
        ref_aut = min(ref, 60.0)
        ref_exc = max(0.0, ref - 60.0)

        auto = row.get("Autogestión", 0.0)
        auto_aut = min(auto, 30.0)
        auto_exc = max(0.0, auto - 30.0)

        feed = row.get("Feedback", 0.0) + row.get("PCA - Feedback", 0.0)
        feed_exc = max(0.0, feed - 30.0) # exceso a fuga

        cur = row.get("Cursos Adicionales", 0.0)
        cur_exc = max(0.0, cur - 60.0) # exceso a fuga

        exc_sem = cdr_exc + ref_exc + auto_exc + feed_exc + cur_exc
        aut_sem = cdr_aut + ref_aut + auto_aut

        filas_sem.append({
            "agente_id": agente_id,
            "semana": semana,
            "exc_sem": exc_sem,
            "aut_sem": aut_sem,
        })

    df_sem = pd.DataFrame(filas_sem)

    # Consolidación por agente
    ag_dias = df_dias.groupby("agente_id").agg(
        tot_con=("t_conectado_min", "sum"),
        tot_exc_dia=("exc_dia", "sum"),
        tot_aut_dia=("aut_dia", "sum"),
        dias_con=("t_conectado_min", "count"),
    )
    ag_sem = df_sem.groupby("agente_id").agg(
        tot_exc_sem=("exc_sem", "sum"),
        tot_aut_sem=("aut_sem", "sum"),
    )

    ag_total = ag_dias.join(ag_sem, how="left").fillna(0.0)
    ag_total["tot_exc"] = ag_total["tot_exc_dia"] + ag_total["tot_exc_sem"]
    ag_total["tot_aut"] = ag_total["tot_aut_dia"] + ag_total["tot_aut_sem"]

    resumen_agente = pd.DataFrame(index=ag_total.index)
    resumen_agente["pct_fuga_con"] = np.where(
        ag_total["tot_con"] > 0,
        (ag_total["tot_exc"] / ag_total["tot_con"] * 100.0).round(1),
        0.0,
    )
    resumen_agente["pct_pausa_aut"] = np.where(
        ag_total["tot_con"] > 0,
        (ag_total["tot_aut"] / ag_total["tot_con"] * 100.0).round(1),
        0.0,
    )
    resumen_agente["pct_productivo"] = np.maximum(
        0.0,
        (100.0 - resumen_agente["pct_fuga_con"] - resumen_agente["pct_pausa_aut"]).round(1),
    )
    resumen_agente["exceso_prom_min"] = np.where(
        ag_total["dias_con"] > 0,
        (ag_total["tot_exc"] / ag_total["dias_con"]).round(1),
        0.0,
    )
    resumen_agente["t_conectado_hrs"] = np.where(
        ag_total["dias_con"] > 0,
        (ag_total["tot_con"] / ag_total["dias_con"] / 60.0).round(1),
        0.0,
    )

    tot_con_g = float(ag_total["tot_con"].sum())
    tot_exc_g = float(ag_total["tot_exc"].sum())
    tot_aut_g = float(ag_total["tot_aut"].sum())
    dias_tot_g = float(ag_total["dias_con"].sum())

    pct_fuga_g = round(tot_exc_g / tot_con_g * 100.0, 1) if tot_con_g > 0 else 0.0
    pct_aut_g = round(tot_aut_g / tot_con_g * 100.0, 1) if tot_con_g > 0 else 0.0
    pct_prod_g = max(0.0, round(100.0 - pct_fuga_g - pct_aut_g, 1)) if tot_con_g > 0 else 0.0
    exceso_prom_g = round(tot_exc_g / dias_tot_g, 1) if dias_tot_g > 0 else 0.0
    t_conectado_hrs_g = round(tot_con_g / dias_tot_g / 60.0, 1) if dias_tot_g > 0 else 0.0

    global_uso = {
        "pct_fuga_con": pct_fuga_g,
        "pct_productivo": pct_prod_g,
        "pct_pausa_aut": pct_aut_g,
        "exceso_prom_min": exceso_prom_g,
        "t_conectado_hrs": t_conectado_hrs_g,
    }

    mapa_diario = {}
    for _, r in df_dias.iterrows():
        aid = r["agente_id"]
        fec = r["fecha"]
        t_c = r["t_conectado_min"]
        e_d = r["exc_dia"]
        p_f = min(100.0, (100.0 * e_d / t_c)) if t_c > 0 else 0.0
        mapa_diario[(aid, fec)] = {
            "pct_fuga": round(float(p_f), 1),
            "t_con_min": round(float(t_c), 1),
            "exceso_min": round(float(e_d), 1),
        }

    return resumen_agente, global_uso, mapa_diario


def render_timeline(df_agente: pd.DataFrame, color_map: dict[str, str]):
    df_agente = df_agente.copy()
    df_agente["inicio"] = pd.to_datetime(df_agente["inicio"])
    df_agente["fin"] = pd.to_datetime(df_agente["fin"]).fillna(df_agente["inicio"] + pd.Timedelta(minutes=1))

    fig = px.timeline(
        df_agente.sort_values("fecha"),
        x_start="inicio",
        x_end="fin",
        y="fecha",
        color="presence_label",
        color_discrete_map=color_map,
        hover_data={"duracion_min": True, "inicio": True, "fin": True, "fecha": False},
    )
    fig.update_yaxes(autorange="reversed", type="category")
    altura = max(220, 32 * df_agente["fecha"].nunique() + 100)
    fig.update_layout(height=altura, showlegend=True, xaxis_title=None, yaxis_title=None, legend_title_text="Estado")
    st.plotly_chart(fig, use_container_width=True)


# ── App ──────────────────────────────────────────────────────────────────

st.markdown(
    "<div style='text-align:center;'>"
    "<h2 style='margin-bottom:0;'>Radar Genesys</h2>"
    "<p style='color:gray; margin-top:0;'>Pausas y Adherencia de Turno</p>"
    "</div>",
    unsafe_allow_html=True,
)


@st.cache_data(ttl=3600)
def cargar_agentes_map_base():
    real_db_path = Path(__file__).parent / DB_PATH
    if not os.path.exists(real_db_path):
        return {}
    conn = sqlite3.connect(real_db_path)
    agentes_db = pd.read_sql(
        "SELECT distinct agente_id, agente, cargo, servicio, jefe_inmediato, coordinador FROM segments", conn
    )
    conn.close()
    if "cargo" in agentes_db.columns:
        agentes_db = agentes_db[agentes_db["cargo"].str.upper().str.contains("ASESOR", na=False)]
    agentes_db = agentes_db.drop_duplicates("agente_id", keep="last")
    return agentes_db.set_index("agente_id").to_dict(orient="index")


tab_historico, tab_vivo = st.tabs(["📊 Análisis Histórico y Adherencia", "🔴 Monitoreo en Vivo (Piso)"])

with tab_historico:
    rango_disponible = cargar_rango_fechas()
    if not rango_disponible:
        st.warning("Todavía no hay datos extraídos. Corre `python extract_presencia.py` primero.")
        st.stop()
    
    fecha_min_disp, fecha_max_disp = rango_disponible
    fecha_max_disp_d = pd.Timestamp(fecha_max_disp).date()
    fecha_min_disp_d = pd.Timestamp(fecha_min_disp).date()
    
    if "desde" not in st.session_state:
        st.session_state["desde"] = max(fecha_min_disp_d, fecha_max_disp_d - pd.Timedelta(days=13))
        st.session_state["hasta"] = fecha_max_disp_d
    
    col_desde, col_hasta, col_coord, col_servicio, col_superv, col_agente = st.columns(6)
    with col_desde:
        st.date_input("Desde", key="desde", min_value=fecha_min_disp_d, max_value=fecha_max_disp_d)
    with col_hasta:
        st.date_input("Hasta", key="hasta", min_value=fecha_min_disp_d, max_value=fecha_max_disp_d)
    
    col_rango1, col_rango2, col_rango3, col_rango4, _, col_caption = st.columns([1, 1, 1, 1, 1, 4])
    def set_rango(dias):
        st.session_state["hasta"] = fecha_max_disp_d
        st.session_state["desde"] = fecha_max_disp_d - pd.Timedelta(days=dias - 1) if dias else fecha_min_disp_d
        if st.session_state["desde"] < fecha_min_disp_d:
            st.session_state["desde"] = fecha_min_disp_d
    
    with col_rango1:
        st.button("7 días", on_click=set_rango, args=(7,), use_container_width=True)
    with col_rango2:
        st.button("14 días", on_click=set_rango, args=(14,), use_container_width=True)
    with col_rango3:
        st.button("30 días", on_click=set_rango, args=(30,), use_container_width=True)
    with col_rango4:
        st.button("Todo", on_click=set_rango, args=(None,), use_container_width=True)
    with col_caption:
        st.markdown(
            f"<p style='text-align:right; color:gray; padding-top:0.5rem;'>Datos disponibles: {fecha_min_disp} → {fecha_max_disp}</p>",
            unsafe_allow_html=True,
        )
    
    fecha_desde = str(st.session_state["desde"])
    fecha_hasta = str(st.session_state["hasta"])
    if fecha_desde > fecha_hasta:
        st.error("La fecha 'Desde' es posterior a 'Hasta'.")
        st.stop()
    
    df = cargar_rango(fecha_desde, fecha_hasta)
    if df.empty:
        st.info("Sin tramos para este rango.")
        st.stop()
    
    FILTROS = [
        ("coord_sel", "coordinador", "Coordinador", col_coord),
        ("servicio_sel", "servicio", "Servicio", col_servicio),
        ("superv_sel", "jefe_inmediato", "Supervisor", col_superv),
        ("agentes_sel", "agente", "Agente", col_agente),
    ]
    
    
    def aplicar_filtros(df: pd.DataFrame, excluir_key: str | None) -> pd.DataFrame:
        vista = df
        for key, columna, _, _ in FILTROS:
            if key == excluir_key:
                continue
            seleccion = st.session_state.get(key, [])
            if seleccion:
                vista = vista[vista[columna].isin(seleccion)]
        return vista
    
    
    for key, columna, label, col in FILTROS:
        opciones = opciones_validas(aplicar_filtros(df, excluir_key=key)[columna])
        seleccion_actual = st.session_state.get(key, [])
        podada = [v for v in seleccion_actual if v in opciones]
        if podada != seleccion_actual:
            st.session_state[key] = podada
        with col:
            st.multiselect(label, opciones, key=key, placeholder="Todos")
    
    vista = aplicar_filtros(df, excluir_key=None)
    if vista.empty:
        st.info("Sin tramos para estos filtros.")
        st.stop()
    
    with st.expander("⚙️ Configuración: Regla de Available / Conectado y Exclusiones", expanded=False):
        c_av1, c_av2 = st.columns([1, 2])
        with c_av1:
            penalizar_available = st.toggle(
                "⚠️ Penalizar Available como Fuga",
                value=True,
                help="Penaliza el tiempo en Available/Conectado para asesores que deberían estar en cola (On Queue).",
            )
        with c_av2:
            todos_los_agentes = sorted(df["agente"].dropna().unique())
            default_excluidos = [a for a in AGENTES_AUTORIZADOS_AVAILABLE_DEFAULT if a in todos_los_agentes]
            excluidos_available = set(
                st.multiselect(
                    "🛡️ Asesores autorizados en Available (Lista blanca / No penalizar):",
                    options=todos_los_agentes,
                    default=default_excluidos,
                    help="Los asesores aquí seleccionados (Backoffice, Equipajes, Cargo, etc.) NO serán penalizados por estar en Available (computa como Productivo).",
                )
            ) if penalizar_available else set()

    turnos_rango = cargar_turnos(fecha_desde, fecha_hasta)
    adherencia_horario = calcular_adherencia_horario(vista, turnos_rango)
    resumen_uso, global_uso, mapa_uso_diario = calcular_uso_turno(
        vista, turnos_rango, penalizar_available=penalizar_available, excluidos_available=excluidos_available
    )
    
    # ── Tarjetas KPI ─────────────────────────────────────────────────────────
    
    globales = cumplimiento_global(vista)
    filtro_kpi_hist = st.session_state.get("hist_filtro_kpi", None)
    kpi_cols = st.columns(4)

    # 1. Tarjeta Adherencia Horario
    with kpi_cols[0]:
        horario_valido = adherencia_horario["horario"].dropna()
        if horario_valido.empty:
            pct_horario_txt, color_horario, subtitulo_horario = "—", "gray", "Sin turnos cargados en el rango"
        else:
            pct_horario = round(horario_valido.mean(), 1)
            pct_horario_txt = f"{pct_horario}%"
            color_horario = color_pct(pct_horario)
            subtitulo_horario = f"{len(horario_valido)} agentes con turno"

        es_act_h = (filtro_kpi_hist == "horario")
        borde_h = f"border: 2px solid {color_horario}; box-shadow: 0 0 10px {color_horario}44; background: #fff;" if es_act_h else "background: #f7f7f7;"
        tag_h = "<span style='float:right; font-size:10px; background:#185fa5; color:#fff; padding:1px 6px; border-radius:8px;'>✓ Filtrando</span>" if es_act_h else ""

        st.markdown(
            f"""
            <div style="{borde_h} border-radius:10px; padding:12px 14px; margin-bottom:4px; transition: all 0.2s;">
                <p style="color:gray; font-size:13px; margin:0;">Adherencia Horario {tag_h}</p>
                <p style="color:{color_horario}; font-size:28px; font-weight:700; margin:2px 0;">{pct_horario_txt}</p>
                <p style="color:gray; font-size:12px; margin:0;">{subtitulo_horario}</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        btn_txt_h = "✖ Quitar filtro" if es_act_h else "🔍 Filtrar <90%"
        if st.button(btn_txt_h, key="btn_hist_kpi_horario", width="stretch", type="primary" if es_act_h else "secondary"):
            st.session_state["hist_filtro_kpi"] = None if es_act_h else "horario"
            st.rerun()

    # 2. Tarjeta % Fuga
    with kpi_cols[1]:
        pct_fuga = global_uso["pct_fuga_con"]
        min_exceso = global_uso["exceso_prom_min"]
        color_f = color_fuga(pct_fuga)
        es_act_f = (filtro_kpi_hist == "fuga")
        borde_f = f"border: 2px solid {color_f}; box-shadow: 0 0 10px {color_f}44; background: #fff;" if es_act_f else "background: #f7f7f7;"
        tag_f = "<span style='float:right; font-size:10px; background:#b3261e; color:#fff; padding:1px 6px; border-radius:8px;'>✓ Filtrando</span>" if es_act_f else ""

        st.markdown(
            f"""
            <div style="{borde_f} border-radius:10px; padding:12px 14px; margin-bottom:4px; transition: all 0.2s;">
                <p style="color:gray; font-size:13px; margin:0;">% Fuga en Conexión {tag_f}</p>
                <p style="color:{color_f}; font-size:28px; font-weight:700; margin:2px 0;">{pct_fuga:.1f}%</p>
                <p style="color:gray; font-size:12px; margin:0;">Exceso prom: {min_exceso:.1f} min/día</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        btn_txt_f = "✖ Quitar filtro" if es_act_f else "🔍 Filtrar con Fuga"
        if st.button(btn_txt_f, key="btn_hist_kpi_fuga", width="stretch", type="primary" if es_act_f else "secondary"):
            st.session_state["hist_filtro_kpi"] = None if es_act_f else "fuga"
            st.rerun()

    # 3. Tarjetas de Pausas (Descanso, Pre Pausa, Baño, Diálogo, Lunch, CDR)
    for i, card in enumerate(CARDS):
        valor = globales[card["key"]]
        if card["tipo"] == "conteo":
            valor_txt, color_valor = f"{valor}", "#378ADD"
        else:
            valor_txt, color_valor = f"{valor}%", color_pct(valor)

        ckey = card["key"]
        es_act_c = (filtro_kpi_hist == ckey)
        borde_c = f"border: 2px solid {color_valor}; box-shadow: 0 0 10px {color_valor}44; background: #fff;" if es_act_c else "background: #f7f7f7;"
        tag_c = "<span style='float:right; font-size:10px; background:#185fa5; color:#fff; padding:1px 6px; border-radius:8px;'>✓ Filtrando</span>" if es_act_c else ""

        with kpi_cols[(i + 2) % 4]:
            st.markdown(
                f"""
                <div style="{borde_c} border-radius:10px; padding:12px 14px; margin-bottom:4px; transition: all 0.2s;">
                    <p style="color:gray; font-size:13px; margin:0;">{card['label']} {tag_c}</p>
                    <p style="color:{color_valor}; font-size:28px; font-weight:700; margin:2px 0;">{valor_txt}</p>
                    <p style="color:gray; font-size:12px; margin:0;">{meta_texto(card)}</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
            btn_txt_c = "✖ Quitar" if es_act_c else ("🔍 Filtrar uso" if card["tipo"] == "conteo" else "🔍 Filtrar <100%")
            if st.button(btn_txt_c, key=f"btn_hist_kpi_{ckey}", width="stretch", type="primary" if es_act_c else "secondary"):
                st.session_state["hist_filtro_kpi"] = None if es_act_c else ckey
                st.rerun()
    
    # ── Tabla por agente ───────────────────────────────────────────────────

    # ── Módulo Comparativo: Uso de Turno vs. % Pausas sobre Conexión ────────
    with st.expander("⚖️ Distribución del 100% del Tiempo Conectado (Balance de Jornada)", expanded=False):
        st.markdown(
            """
            **Composición del 100% del tiempo conectado:**  
            `% Productivo` + `% Pausas Autorizadas` representan la porción válida y legítima. El **% Fuga** es el faltante para completar el 100%.
            """
        )
        c_kpi1, c_kpi2, c_kpi3, c_kpi4 = st.columns(4)
        c_kpi1.metric("1. % Fuga (Excesos)", f"{global_uso['pct_fuga_con']:.1f}%", f"Exceso: {global_uso['exceso_prom_min']:.1f} m/día", delta_color="inverse")
        c_kpi2.metric("2. % Productivo", f"{global_uso['pct_productivo']:.1f}%", "Atención y gestión", delta_color="off")
        c_kpi3.metric("3. % Pausas Autorizadas", f"{global_uso['pct_pausa_aut']:.1f}%", "Dentro de metas", delta_color="off")
        c_kpi4.metric("4. Conectado Promedio", f"{global_uso['t_conectado_hrs']:.1f} h/día", "Base de conexión")

        prod_g, aut_g, exc_g, con_g, _, _, _ = calcular_componentes_100(
            vista, "", penalizar_available=penalizar_available, excluidos_available=excluidos_available
        )
        render_distribucion_100(prod_g, aut_g, exc_g, con_g, titulo="Distribución Global del Tiempo Conectado (Filtros actuales)")

    micro_por_agente = (
        vista[
            (~vista["presence_label"].isin(ESTADOS_SISTEMA)) &
            (vista["duracion_min"] <= 0.25)
        ]
        .groupby("agente_id")
        .size()
        .rename("micro_estados")
    )

    tabla = (
        calcular_cumplimiento(vista)
        .join(adherencia_horario[["horario", "dias_con_turno"]])
        .join(resumen_uso[["pct_fuga_con", "pct_productivo", "pct_pausa_aut", "exceso_prom_min", "t_conectado_hrs"]])
        .join(micro_por_agente)
    )
    tabla["micro_estados"] = tabla["micro_estados"].fillna(0).astype(int)
    n_agentes = len(tabla)
    st.caption(f"{n_agentes} agentes · ordenados por mayor % de fuga · % Fuga + % Productivo + % Pausas Aut. = 100% · clic en una fila para ver el detalle")

    columnas_mostrar = [
        "agente", "coordinador", "servicio", "jefe_inmediato",
        "pct_fuga_con", "pct_productivo", "pct_pausa_aut", "exceso_prom_min", "t_conectado_hrs",
        "micro_estados", "horario"
    ] + [c["key"] for c in CARDS]

    tabla_mostrar = (
        tabla.reset_index()[["agente_id"] + columnas_mostrar]
        .rename(
            columns={
                "agente": "Agente", "coordinador": "Coordinador", "servicio": "Servicio",
                "jefe_inmediato": "Supervisor",
                "pct_fuga_con": "% Fuga",
                "pct_productivo": "% Productivo",
                "pct_pausa_aut": "% Pausas Aut.",
                "exceso_prom_min": "Exceso (min)",
                "t_conectado_hrs": "Conectado (h)",
                "micro_estados": "Micro (≤15s)", "horario": "Horario",
                **{c["key"]: c["label"] for c in CARDS},
            }
        )
        .sort_values(by=["% Fuga", "Exceso (min)"], ascending=[False, False])
    )

    # Filtrado interactivo activado desde tarjetas KPI
    filtro_kpi_hist = st.session_state.get("hist_filtro_kpi", None)
    if filtro_kpi_hist:
        nombre_filtro_kpi = ""
        if filtro_kpi_hist == "horario":
            nombre_filtro_kpi = "Adherencia Horario (< 90%)"
            tabla_mostrar = tabla_mostrar[(tabla_mostrar["Horario"].notna()) & (tabla_mostrar["Horario"] < 90.0)].sort_values(by="Horario", ascending=True)
        elif filtro_kpi_hist == "fuga":
            nombre_filtro_kpi = "% Fuga en Conexión (> 0%)"
            tabla_mostrar = tabla_mostrar[tabla_mostrar["% Fuga"] > 0.0].sort_values(by=["% Fuga", "Exceso (min)"], ascending=[False, False])
        else:
            for card in CARDS:
                if card["key"] == filtro_kpi_hist:
                    c_label = card["label"]
                    nombre_filtro_kpi = f"{c_label} (Uso o Exceso)"
                    if card["tipo"] == "conteo":
                        tabla_mostrar = tabla_mostrar[tabla_mostrar[c_label] > 0]
                    else:
                        tabla_mostrar = tabla_mostrar[(tabla_mostrar[c_label].notna()) & (tabla_mostrar[c_label] < 100.0)].sort_values(by=c_label, ascending=True)
                    break

        st.info(
            f"🔍 **Filtro de Tarjeta Activo:** Mostrando **{len(tabla_mostrar)} asesores** filtrados por **{nombre_filtro_kpi}**. "
            "Haz clic en el botón de la tarjeta activa para ver todos."
        )

    pct_cols = ["% Fuga", "% Productivo", "% Pausas Aut.", "Horario"] + [c["label"] for c in CARDS if c["tipo"] != "conteo"]
    conteo_cols = ["Micro (≤15s)"] + [c["label"] for c in CARDS if c["tipo"] == "conteo"]


    def estilo_pct(val):
        if pd.isna(val):
            return ""
        return f"background-color: {color_pct(val)}22; color: {color_pct(val)}; font-weight: 600;"


    def estilo_fuga_celda(val):
        if pd.isna(val):
            return ""
        return f"background-color: {color_fuga(val)}22; color: {color_fuga(val)}; font-weight: 600;"


    styler = (
        tabla_mostrar.drop(columns=["agente_id"])
        .style.map(estilo_pct, subset=["Horario"] + [c["label"] for c in CARDS if c["tipo"] != "conteo"])
        .map(estilo_fuga_celda, subset=["% Fuga"])
        .map(estilo_prod, subset=["% Productivo"])
        .map(estilo_aut, subset=["% Pausas Aut."])
        .map(estilo_micro, subset=["Micro (≤15s)"])
    )

    column_config = {c: st.column_config.NumberColumn(c, format="%.1f%%") for c in pct_cols}
    column_config.update({c: st.column_config.NumberColumn(c, format="%d") for c in conteo_cols})
    column_config["% Fuga"] = st.column_config.NumberColumn("% Fuga", format="%.1f%%", help="Pausas no productivas o excesos sobre el tiempo conectado (faltante para el 100%)")
    column_config["% Productivo"] = st.column_config.NumberColumn("% Productivo", format="%.1f%%", help="% del tiempo conectado en atención y gestión operativa")
    column_config["% Pausas Aut."] = st.column_config.NumberColumn("% Pausas Aut.", format="%.1f%%", help="% del tiempo conectado en pausas reglamentarias dentro de meta")
    column_config["Exceso (min)"] = st.column_config.NumberColumn("Exceso (min)", format="%.1f m", help="Minutos diarios promedio de exceso sobre metas de pausas")
    column_config["Conectado (h)"] = st.column_config.NumberColumn("Conectado (h)", format="%.1f h", help="Horas promedio de conexión diaria")
    column_config["Micro (≤15s)"] = st.column_config.NumberColumn(
        "Micro (≤15s)",
        format="%d",
        help="Tramos en pausas o gestiones con duración ≤ 15 seg. Ítem a revisar por posibles refrescos de cola o alternancia operativa.",
    )

    evento = st.dataframe(
        styler,
        column_config=column_config,
        use_container_width=True,
        hide_index=True,
        on_select="rerun",
        selection_mode="single-row",
        key="tabla_agentes",
    )
    
    # ── Exportar Excel ───────────────────────────────────────────────────────
    
    presencias_de_interes = [p for c in CARDS for p in c["presence"]]
    conteo_general = (
        vista[vista["presence_label"].isin(presencias_de_interes)]
        .groupby(["agente", "presence_label"])
        .size()
        .unstack(fill_value=0)
        .rename(columns=lambda p: f"{p} (veces)")
    )
    
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        hoja_radar = tabla_mostrar.drop(columns=["agente_id"])
        hoja_radar.to_excel(writer, index=False, sheet_name="Radar Genesys")
        conteo_general.reset_index().to_excel(writer, index=False, sheet_name="Cantidad de pausas")
    
        # Excel muestra los floats con toda su precision binaria (ej. 71.79999999999999)
        # si no se fija un formato de celda explicito - el .round(1) de pandas no alcanza.
        ws = writer.sheets["Radar Genesys"]
        for nombre_col in pct_cols + conteo_cols:
            idx = hoja_radar.columns.get_loc(nombre_col) + 1
            letra = get_column_letter(idx)
            formato = '0.0"%"' if nombre_col in pct_cols else "0"
            for celda in ws[f"{letra}2:{letra}{ws.max_row}"]:
                celda[0].number_format = formato
    st.download_button(
        "📥 Exportar Excel",
        buffer.getvalue(),
        file_name=f"radar_genesys_{fecha_desde}_a_{fecha_hasta}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    
    # ── Detalle del agente seleccionado ─────────────────────────────────────
    
    filas_sel = evento.selection.rows if evento and evento.selection else []
    if filas_sel:
        agente_id_sel = tabla_mostrar.iloc[filas_sel[0]]["agente_id"]
        fila_sel = tabla.loc[agente_id_sel]
    
        st.divider()
        st.subheader(f"Detalle — {fila_sel['agente']}")
        st.caption(f"{fila_sel['servicio']} · Supervisor: {fila_sel['jefe_inmediato']}")

        c_da1, c_da2, c_da3, c_da4 = st.columns(4)
        c_da1.metric("% Fuga s/ Conexión", f"{fila_sel['pct_fuga_con']:.1f}%", help="Minutos de pausas no productivas o excesos ÷ tiempo conectado (faltante para el 100%)")
        c_da2.metric("% Productivo", f"{fila_sel['pct_productivo']:.1f}%", help="% del tiempo conectado en atención y gestión operativa")
        c_da3.metric("% Pausas Autorizadas", f"{fila_sel['pct_pausa_aut']:.1f}%", help="% del tiempo conectado en pausas reglamentarias dentro de meta")
        c_da4.metric("Conectado Diario Prom.", f"{fila_sel['t_conectado_hrs']:.1f} h/día", f"Exceso: {fila_sel['exceso_prom_min']:.1f} m/día")

        df_agente = vista[vista["agente_id"] == agente_id_sel]
        if fila_sel["agente"] in excluidos_available:
            st.info("🛡️ **Asesor en Lista Blanca:** Autorizado para operar en `Available` (computa como tiempo productivo).")
        prod_a, aut_a, exc_a, con_a, det_pa, det_aa, det_ea = calcular_componentes_100(
            df_agente, fila_sel["servicio"], penalizar_available=penalizar_available, excluidos_available=excluidos_available
        )
        render_distribucion_100(prod_a, aut_a, exc_a, con_a, titulo=f"Distribución del 100% Conectado — {fila_sel['agente']}")

        with st.expander("🔍 Desglose detallado del 100% (Productivo vs. Pausas)", expanded=False):
            col_dp1, col_dp2, col_dp3 = st.columns(3)
            with col_dp1:
                pct_p = (prod_a / con_a * 100) if con_a > 0 else 0.0
                st.markdown(f"**🟢 Operación Productiva ({pct_p:.1f}%)**")
                for k, v in sorted(det_pa.items(), key=lambda x: -x[1]):
                    st.write(f"- {k}: **{v:.1f} min** ({(v/con_a*100):.1f}%)")
            with col_dp2:
                pct_a = (aut_a / con_a * 100) if con_a > 0 else 0.0
                st.markdown(f"**🔵 Pausas Autorizadas ({pct_a:.1f}%)**")
                for k, v in sorted(det_aa.items(), key=lambda x: -x[1]):
                    if v > 0:
                        st.write(f"- {k}: **{v:.1f} min** ({(v/con_a*100):.1f}%)")
                if aut_a == 0:
                    st.caption("Sin pausas autorizadas registradas")
            with col_dp3:
                pct_e = (exc_a / con_a * 100) if con_a > 0 else 0.0
                st.markdown(f"**🔴 Excesos / Fuga ({pct_e:.1f}%)**")
                for k, v in sorted(det_ea.items(), key=lambda x: -x[1]):
                    if v > 0:
                        st.write(f"- {k}: **{v:.1f} min** ({(v/con_a*100):.1f}%)")
                if exc_a == 0:
                    st.caption("✅ Sin excesos ni pausas no autorizadas")
    
        vista_servicio = vista[vista["servicio"] == fila_sel["servicio"]]
        n_agentes_servicio = vista_servicio["agente_id"].nunique()
        df_agente = vista[vista["agente_id"] == agente_id_sel]
        bp_agente = fila_sel["agente"].split(" - ")[0].strip()
        conteo_por_estado = df_agente.groupby("presence_label").size()
    
        # Micro-estados del asesor (tramos auxiliares <= 15 seg)
        df_agente_valid = df_agente[~df_agente["presence_label"].isin(ESTADOS_SISTEMA)]
        micro_agente = df_agente_valid[df_agente_valid["duracion_min"] <= 0.25]
        conteo_micro = micro_agente.groupby("presence_label").size()
        total_micro = len(micro_agente)
    
        diario_agente = df_agente.groupby(["fecha", "presence_label"])["duracion_min"].sum().reset_index()
        pivot_agente_dia = diario_agente.pivot_table(index="fecha", columns="presence_label", values="duracion_min", fill_value=0)
    
        # Alerta amistosa de auditoría / ítem a revisar
        if total_micro > 0:
            st.info(
                f"🔍 **Ítem a revisar:** Se identificaron **{total_micro} micro-estados (≤ 15 seg)** en el período. "
                f"En servicios mixtos (voz y no-voz, como DT FFP / Trim Team) puede deberse a alternancia operativa entre canales, "
                f"pero se sugiere validar con el asesor para descartar posibles refrescos de cola o evasión de turnos.",
                icon="🔍",
            )
    
        # ── Tabla Maestra Unificada de Pausas y Gestiones ─────────────────
        st.markdown(f"**Uso de Pausas y Gestiones — vs. Metas y Mediana de «{fila_sel['servicio']}»** ({n_agentes_servicio} agentes en el servicio)")
    
        # Formatear duración: si promedio >= 0.1 min -> X.X min, si > 0 pero < 0.1 -> X seg, si 0 -> 0.0 min
        def format_duracion(prom_min, dur_total_min):
            if prom_min >= 0.1:
                return f"{prom_min:.1f} min"
            elif dur_total_min > 0:
                sec = int(round(dur_total_min * 60))
                return f"{sec} seg" if sec > 0 else "< 1 seg"
            else:
                return "0.0 min"
    
        # Duración promedio de turno del agente
        fechas_agente = sorted(pivot_agente_dia.index)
        duraciones_turno = []
        for f in fechas_agente:
            turnos_dia = turnos_rango[(turnos_rango["bp"] == bp_agente) & (turnos_rango["fecha"] == f)]
            if not turnos_dia.empty:
                r = turnos_dia.iloc[0]
                f_dt = pd.Timestamp(r["fecha"])
                ini = f_dt + pd.to_timedelta(r["hora_inicio"])
                fin = f_dt + pd.to_timedelta(r["hora_fin"])
                if fin <= ini:
                    fin += pd.Timedelta(days=1)
                dur = (fin - ini).total_seconds() / 60.0
                if dur > 0:
                    duraciones_turno.append(dur)
        if not duraciones_turno:
            conectado_dia = df_agente[df_agente["presence_label"] != "Offline"].groupby("fecha")["duracion_min"].sum()
            promedio_turno_agente = conectado_dia.mean() if not conectado_dia.empty and conectado_dia.mean() > 0 else 480.0
        else:
            promedio_turno_agente = sum(duraciones_turno) / len(duraciones_turno)
    
        filas_unificadas = []
        tiene_prepausa_serv = servicio_tiene_prepausa(fila_sel["servicio"])
    
        # 1. Pausas reglamentarias con meta
        for c in CARDS:
            veces = int(sum(conteo_por_estado.get(p, 0) for p in c["presence"]))
            micro_c = int(sum(conteo_micro.get(p, 0) for p in c["presence"]))
            dur_total = float(df_agente[df_agente["presence_label"].isin(c["presence"])]["duracion_min"].sum())
            usado_serie = sumar_presencias(pivot_agente_dia, c["presence"])
            prom_min = float(usado_serie.mean()) if len(usado_serie) else 0.0
            pct_turno = (prom_min / promedio_turno_agente) * 100.0 if promedio_turno_agente > 0 else 0.0
    
            if c["unidad"] == "dia":
                if c["key"] == "lunch":
                    meta_txt = "0 min (No aut.)"
                    dif_min = prom_min
                    adh = 100.0 if prom_min == 0 else max(0.0, 100.0 - (prom_min / promedio_turno_agente * 100.0))
                elif c["key"] == "pre_pausa":
                    meta_val = 60 if tiene_prepausa_serv else 0
                    meta_txt = f"{meta_val} min" if tiene_prepausa_serv else "0 min (No aut.)"
                    dif_min = prom_min - meta_val
                    adh = 100.0 if prom_min <= meta_val else (round(100.0 * meta_val / prom_min, 1) if meta_val > 0 else 0.0)
                elif c["key"] == "bano":
                    meta_txt = f"{c['meta']} min"
                    dif_min = prom_min - c["meta"]
                    adh = 100.0 if prom_min <= c["meta"] else round(100.0 * c["meta"] / prom_min, 1)
                else:
                    meta_txt = f"{c['meta']} min"
                    dif_min = prom_min - c["meta"]
                    adh = fila_sel[c["key"]]
            else:
                # CDR semanal
                meta_txt = f"{c['meta']} min/sem"
                dif_min = prom_min - (c["meta"] / 5.0)
                adh = 100.0 if prom_min == 0 else fila_sel[c["key"]]
    
            filas_unificadas.append({
                "Categoría": "Pausa Reglamentaria",
                "Pausa / Estado": c["label"],
                "Veces": veces,
                "Micro (≤15s)": micro_c,
                "Min. promedio/día": format_duracion(prom_min, dur_total),
                "% del Turno": round(pct_turno, 1),
                "Meta / Referencia": meta_txt,
                "Dif. vs Referencia": round(dif_min, 1),
                "Adherencia": f"{adh:.1f}%" if pd.notna(adh) else "—",
                "_adh_num": float(adh) if pd.notna(adh) else 999.0,
            })
    
        # 2. Gestiones operativas sin meta (comparadas contra la mediana del servicio)
        presencias_con_meta = {p for c in CARDS for p in c["presence"]}
        otras_presencias = sorted(
            set(vista_servicio["presence_label"].unique()) - presencias_con_meta - ESTADOS_SISTEMA
        )
    
        if otras_presencias and not vista_servicio.empty:
            diario_servicio = vista_servicio.groupby(["agente_id", "fecha", "presence_label"])["duracion_min"].sum().reset_index()
            pivot_servicio = diario_servicio.pivot_table(index=["agente_id", "fecha"], columns="presence_label", values="duracion_min", fill_value=0)
            promedio_por_agente = pivot_servicio.groupby("agente_id").mean()
    
            for label in otras_presencias:
                if label not in promedio_por_agente.columns:
                    continue
                prom_agente = float(promedio_por_agente.loc[agente_id_sel, label]) if agente_id_sel in promedio_por_agente.index else 0.0
                dur_total = float(df_agente[df_agente["presence_label"] == label]["duracion_min"].sum())
                mediana_servicio = float(promedio_por_agente[label].median())
                pct_turno = (prom_agente / promedio_turno_agente) * 100.0 if promedio_turno_agente > 0 else 0.0
                dif_equipo = prom_agente - mediana_servicio
                micro_l = int(conteo_micro.get(label, 0))
    
                filas_unificadas.append({
                    "Categoría": "Gestión Operativa",
                    "Pausa / Estado": label,
                    "Veces": int(conteo_por_estado.get(label, 0)),
                    "Micro (≤15s)": micro_l,
                    "Min. promedio/día": format_duracion(prom_agente, dur_total),
                    "% del Turno": round(pct_turno, 1),
                    "Meta / Referencia": f"Mediana: {mediana_servicio:.1f} min",
                    "Dif. vs Referencia": round(dif_equipo, 1),
                    "Adherencia": "—",
                    "_adh_num": 999.0,
                })
    
        tabla_unificada = pd.DataFrame(filas_unificadas)
        if not tabla_unificada.empty:
            df_meta = tabla_unificada[tabla_unificada["Categoría"] == "Pausa Reglamentaria"].sort_values(
                by="_adh_num", ascending=True
            )
            df_gestion = tabla_unificada[tabla_unificada["Categoría"] == "Gestión Operativa"].sort_values(
                by="Dif. vs Referencia", ascending=False
            )
            tabla_unificada = pd.concat([df_meta, df_gestion], ignore_index=True).drop(columns=["_adh_num"])
    
        def estilo_dif(val):
            if pd.isna(val):
                return ""
            if val > 15:
                color = "#e24b4a"
            elif val > 5:
                color = "#eda100"
            else:
                color = "#1baf7a"
            return f"background-color: {color}22; color: {color}; font-weight: 600;"
    
        def estilo_adh(val):
            if not val or val == "—":
                return ""
            try:
                num = float(str(val).replace("%", "").strip())
                return f"background-color: {color_pct(num)}22; color: {color_pct(num)}; font-weight: 600;"
            except Exception:
                return ""
    
        styler_unificado = (
            tabla_unificada.style
            .map(estilo_dif, subset=["Dif. vs Referencia"])
            .map(estilo_micro, subset=["Micro (≤15s)"])
            .map(estilo_adh, subset=["Adherencia"])
        )
    
        st.dataframe(
            styler_unificado,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Categoría": st.column_config.TextColumn("Categoría"),
                "Pausa / Estado": st.column_config.TextColumn("Pausa / Estado"),
                "Veces": st.column_config.NumberColumn("Veces", format="%d"),
                "Micro (≤15s)": st.column_config.NumberColumn(
                    "Micro (≤15s)",
                    format="%d",
                    help="Tramos con duración ≤ 15 seg. Ítem a revisar por posibles refrescos de cola o alternancia operativa.",
                ),
                "Min. promedio/día": st.column_config.TextColumn("Min. promedio/día"),
                "% del Turno": st.column_config.NumberColumn("% del Turno", format="%.1f%%"),
                "Meta / Referencia": st.column_config.TextColumn("Meta / Referencia"),
                "Dif. vs Referencia": st.column_config.NumberColumn("Dif. vs Referencia", format="%+.1f min"),
                "Adherencia": st.column_config.TextColumn("Adherencia"),
            },
        )
    
        color_map = cargar_color_map()
        render_timeline(df_agente, color_map)
    
        # ── Detalle diario (score por dia) ──────────────────────────────
        fechas_agente = sorted(pivot_agente_dia.index)
        st.subheader(f"Detalle diario ({len(fechas_agente)} días)")
    
        filas_diarias = []
        for fecha in fechas_agente:
            df_dia = df_agente[df_agente["fecha"] == fecha]
            turnos_dia = turnos_rango[(turnos_rango["bp"] == bp_agente) & (turnos_rango["fecha"] == fecha)]
            adh_dia = calcular_adherencia_horario(df_dia, turnos_dia)
            horario_dia = adh_dia["horario"].iloc[0] if not adh_dia.empty else float("nan")
            info_dia = mapa_uso_diario.get((agente_id_sel, fecha), {})
            fuga_dia = info_dia.get("pct_fuga", float("nan")) if isinstance(info_dia, dict) else float("nan")
            prod_dia = info_dia.get("pct_prod", float("nan")) if isinstance(info_dia, dict) else float("nan")
            aut_dia = info_dia.get("pct_aut", float("nan")) if isinstance(info_dia, dict) else float("nan")
            exc_dia = info_dia.get("exceso_min", float("nan")) if isinstance(info_dia, dict) else float("nan")

            fila = {
                "Fecha": fecha,
                "Novedad": "—",
                "% Fuga": fuga_dia,
                "% Productivo": prod_dia,
                "% Pausas Aut.": aut_dia,
                "Exceso (min)": exc_dia,
                "Horario": horario_dia,
            }
            fila_valores = pivot_agente_dia.loc[[fecha]]
            for c in CARDS:
                usado_dia = sumar_presencias(fila_valores, c["presence"]).iloc[0]
                if c["tipo"] == "conteo":
                    fila[c["label"]] = int(df_dia["presence_label"].isin(c["presence"]).sum())
                else:
                    fila[c["label"]] = score_graduado(pd.Series([usado_dia]), c["meta"]).iloc[0]
            filas_diarias.append(fila)

        tabla_diaria = pd.DataFrame(filas_diarias).sort_values(by=["% Fuga", "Exceso (min)"], ascending=[False, False])
        pct_cols_diario = ["% Fuga", "% Productivo", "% Pausas Aut.", "Horario"] + [c["label"] for c in CARDS if c["tipo"] != "conteo"]
        conteo_cols_diario = [c["label"] for c in CARDS if c["tipo"] == "conteo"]
        styler_diario = (
            tabla_diaria.style.map(estilo_pct, subset=["Horario"] + [c["label"] for c in CARDS if c["tipo"] != "conteo"])
            .map(estilo_fuga_celda, subset=["% Fuga"])
            .map(estilo_prod, subset=["% Productivo"])
            .map(estilo_aut, subset=["% Pausas Aut."])
        )
        col_config_diario = {c: st.column_config.NumberColumn(c, format="%.1f%%") for c in pct_cols_diario}
        col_config_diario["Exceso (min)"] = st.column_config.NumberColumn("Exceso (min)", format="%.1f m")
        col_config_diario.update({c: st.column_config.NumberColumn(c, format="%d") for c in conteo_cols_diario})
        st.dataframe(
            styler_diario,
            use_container_width=True,
            hide_index=True,
            column_config=col_config_diario,
        )

        st.subheader("Detalle de cada ítem")
        st.caption("Cada tramo registrado en Genesys, para todos los días del rango filtrado.")
        detalle_diario = df_agente[["fecha", "presence_label", "inicio", "fin", "duracion_min"]].copy()
        detalle_diario["inicio"] = pd.to_datetime(detalle_diario["inicio"]).dt.strftime("%H:%M:%S")
        detalle_diario["fin"] = pd.to_datetime(detalle_diario["fin"]).dt.strftime("%H:%M:%S")
        detalle_diario["duracion_min"] = detalle_diario["duracion_min"].round(1)
        detalle_diario = detalle_diario.sort_values(["fecha", "inicio"]).rename(
            columns={
                "fecha": "Fecha", "presence_label": "Estado", "inicio": "Hora inicio",
                "fin": "Hora fin", "duracion_min": "Duración (min)",
            }
        )
        st.dataframe(
            detalle_diario,
            use_container_width=True,
            hide_index=True,
            column_config={"Duración (min)": st.column_config.NumberColumn("Duración (min)", format="%.1f")},
        )


with tab_vivo:
    render_tab_en_vivo(cargar_agentes_map_base())
